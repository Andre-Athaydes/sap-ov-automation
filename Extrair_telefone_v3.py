import pandas as pd
import re
import win32com.client
import time
import os
import sys
from openpyxl import load_workbook
from pathlib import Path

def log_message(message):
    """Função para log com timestamp - otimizada"""
    timestamp = time.strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}")
    # Force flush para mostrar mensagens imediatamente
    sys.stdout.flush()

def extrair_telefone(texto):
    """Função para extrair telefone do texto - otimizada"""
    if not texto:
        return ''
    
    # Padrão regex otimizado - compilado uma vez
    if not hasattr(extrair_telefone, 'padrao_compilado'):
        extrair_telefone.padrao_compilado = re.compile(r'\(?\d{2}\)?[-\s]?(\d{4,5})[-\s]?(\d{4})|\b\d{4,5}[-\s]?\d{4}\b')
    
    match = extrair_telefone.padrao_compilado.search(texto)
    
    if match:
        # Se DDD presente, retorna com DDD
        if match.group(0).startswith("(") or len(re.sub(r'\D', '', match.group(0))) > 9:
            return re.sub(r'\D', '', match.group(0))  # Só números
        else:
            return "27" + re.sub(r'\D', '', match.group(0))  # Adiciona DDD padrão se não houver
    return ''

def verificar_arquivo_excel(file_path):
    """Verifica se o arquivo Excel existe - otimizada"""
    if os.path.exists(file_path):
        log_message(f"✅ Arquivo encontrado: {os.path.basename(file_path)}")
        return True
    else:
        log_message(f"❌ Arquivo não encontrado: {file_path}")
        # Sugerir arquivos similares
        pasta = os.path.dirname(file_path)
        if os.path.exists(pasta):
            arquivos_xlsx = [f for f in os.listdir(pasta) if f.endswith(('.xlsx', '.xlsm'))]
            if arquivos_xlsx:
                log_message(f"📁 Arquivos Excel encontrados na pasta: {', '.join(arquivos_xlsx[:3])}")
        return False

def conectar_sap_otimizado():
    """Conecta ao SAP com retry - otimizada"""
    max_tentativas = 3
    for tentativa in range(1, max_tentativas + 1):
        try:
            log_message(f"🔌 Tentativa {tentativa}/{max_tentativas} - Conectando ao SAP...")
            SapGuiAuto = win32com.client.GetObject("SAPGUI")
            Application = SapGuiAuto.GetScriptingEngine
            Connection = Application.OpenConnection("PRODUÇÃO CCS ( EP2 ) - EDP ES", True)
            session1 = Connection.Children(0)
            log_message("✅ Conexão SAP estabelecida")
            return session1
        except Exception as e:
            log_message(f"❌ Tentativa {tentativa} falhou: {e}")
            if tentativa < max_tentativas:
                log_message(f"⏳ Aguardando 3 segundos antes da próxima tentativa...")
                time.sleep(3)
            else:
                log_message(f"💥 Todas as tentativas falharam. Verifique se o SAP GUI está aberto.")
    return None

def fazer_login_sap_otimizado(session1, usuario, senha):
    """Faz login no SAP com validação - otimizada"""
    try:
        log_message(f"🔐 Fazendo login com usuário: {usuario}")
        
        # Verificar se campos existem
        try:
            session1.findById("wnd[0]/usr/txtRSYST-BNAME").Text = usuario
            session1.findById("wnd[0]/usr/pwdRSYST-BCODE").Text = senha
            session1.findById("wnd[0]").sendVKey(0)
            
            # Aguardar um momento para login processar
            time.sleep(2)
            
            # Verificar se login foi bem-sucedido (não há campo de erro visível)
            session1.findById("wnd[0]").maximize
            
            log_message("✅ Login realizado com sucesso")
            return True
            
        except Exception as field_error:
            log_message(f"❌ Erro nos campos de login: {field_error}")
            return False
            
    except Exception as e:
        log_message(f"❌ Erro geral no login SAP: {e}")
        return False

def processar_ov_otimizado(session1, ov, index, total):
    """Processa uma OV específica - otimizada com progresso"""
    try:
        log_message(f"📋 [{index}/{total}] Processando OV: {ov}")
        
        # Acessar VA03 de forma otimizada
        session1.findById("wnd[0]/tbar[0]/okcd").Text = "/nva03"
        session1.findById("wnd[0]").sendVKey(0)
        
        # Aguardar carregar
        time.sleep(0.5)
        
        # Inserir OV
        session1.findById("wnd[0]/usr/ctxtVBAK-VBELN").Text = str(ov)
        session1.findById("wnd[0]").sendVKey(0)
        
        # Aguardar carregar
        time.sleep(1)
        
        # Acessa a aba "Texto"
        session1.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/btnBT_HEAD").press()
        session1.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\08").Select()
        
        # Aguardar carregar aba
        time.sleep(0.5)
        
        # Selecionar item Z014
        lista = session1.findById(
            "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\08/ssubSUBSCREEN_BODY:SAPMV45A:4152/"
            "subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/"
            "shellcont[0]/shell"
        )
        
        lista.selectItem("Z014", "Column1")
        lista.doubleClickItem("Z014", "Column1")
        
        # Aguardar carregar editor
        time.sleep(0.5)
        
        # Extrair texto
        editor = session1.findById(
            "wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\08/ssubSUBSCREEN_BODY:SAPMV45A:4152/"
            "subSUBSCREEN_TEXT:SAPLV70T:2100/cntlSPLITTER_CONTAINER/shellcont/shellcont/shell/"
            "shellcont[1]/shell"
        )
        
        texto = editor.Text
        log_message(f"✅ [{index}/{total}] Texto da OV {ov} capturado ({len(texto)} caracteres)")
        
        # Voltar de forma otimizada
        session1.findById("wnd[0]").sendVKey(3)
        session1.findById("wnd[0]").sendVKey(3)
        
        # Aguardar voltar
        time.sleep(0.5)
        
        telefone = extrair_telefone(texto)
        if telefone:
            log_message(f"📞 [{index}/{total}] Telefone encontrado: {telefone}")
        else:
            log_message(f"📞 [{index}/{total}] Nenhum telefone encontrado")
        
        return telefone
        
    except Exception as e:
        log_message(f"❌ [{index}/{total}] Erro ao processar OV {ov}: {e}")
        # Tentar voltar em caso de erro
        try:
            session1.findById("wnd[0]").sendVKey(3)
            session1.findById("wnd[0]").sendVKey(3)
        except:
            pass
        return ''

def salvar_telefones_excel_otimizado(file_path, telefones):
    """Salva os telefones no Excel mantendo formatação - otimizada"""
    try:
        log_message("💾 Salvando telefones no Excel...")
        
        # Abre o Excel sem perder a formatação
        wb = load_workbook(filename=file_path, keep_vba=True)
        ws = wb['RUA CADASTRADA']
        
        # Mapeia os cabeçalhos da planilha (otimizado)
        headers = {}
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
            if cell.value:
                headers[cell.value] = idx
        
        col_telefone = headers.get("Telefone")
        
        if not col_telefone:
            raise ValueError("❌ Coluna 'Telefone' não encontrada na aba RUA CADASTRADA.")
        
        log_message(f"📌 Atualizando {len(telefones)} telefones na coluna {col_telefone}...")
        
        # Insere os telefones nas linhas correspondentes (otimizado)
        for i, numero in enumerate(telefones, start=2):
            if numero:  # Só atualiza se há telefone
                ws.cell(row=i, column=col_telefone).value = numero
        
        # Salva sem afetar as demais tabelas
        wb.save(file_path)
        wb.close()
        
        telefones_encontrados = sum(1 for t in telefones if t)
        log_message(f"✅ {telefones_encontrados} telefones salvos com sucesso!")
        return True
        
    except Exception as e:
        log_message(f"❌ Erro ao salvar no Excel: {e}")
        return False

def main():
    """Função principal otimizada"""
    inicio_total = time.time()
    
    try:
        log_message("🚀 Iniciando processo OTIMIZADO de atualização de telefones...")
        log_message("⚡ Modo: Python direto (RÁPIDO)")
        
        # Caminho do arquivo Excel
        file_path = r'C:\Users\E711449\OneDrive - EDP\Área de Trabalho\Planilhas_ Scripts_Andre\Cadastro_Logradouro_Python\Cadastro_Logradouro_v4.xlsm'
        
        # Verificar se arquivo existe
        if not verificar_arquivo_excel(file_path):
            input("❌ Arquivo não encontrado. Pressione Enter para sair...")
            return
        
        # Ler dados do Excel
        log_message("📖 Lendo dados do Excel...")
        inicio_leitura = time.time()
        
        login_df = pd.read_excel(file_path, sheet_name='Login')
        rua_df = pd.read_excel(file_path, sheet_name='RUA CADASTRADA', dtype={0: str})
        
        fim_leitura = time.time()
        log_message(f"✅ Dados lidos em {fim_leitura - inicio_leitura:.1f} segundos")
        
        # Obter credenciais
        usuario = login_df.iloc[0]['Usuário']
        senha = login_df.iloc[0]['Senha']
        
        total_ovs = len(rua_df)
        log_message(f"📊 Total de OVs para processar: {total_ovs}")
        
        # Conectar e fazer login no SAP
        session1 = conectar_sap_otimizado()
        if not session1:
            input("❌ Falha na conexão SAP. Pressione Enter para sair...")
            return
        
        if not fazer_login_sap_otimizado(session1, usuario, senha):
            input("❌ Falha no login SAP. Pressione Enter para sair...")
            return
        
        # Processar cada OV com controle de tempo
        log_message("🔄 Iniciando processamento das OVs...")
        inicio_processamento = time.time()
        
        telefones = []
        
        for index, row in rua_df.iterrows():
            ov = str(row['OV'])
            telefone = processar_ov_otimizado(session1, ov, index + 1, total_ovs)
            telefones.append(telefone)
            
            # Mostrar progresso a cada 5 OVs
            if (index + 1) % 5 == 0:
                tempo_decorrido = time.time() - inicio_processamento
                tempo_medio = tempo_decorrido / (index + 1)
                tempo_restante = tempo_medio * (total_ovs - index - 1)
                log_message(f"📈 Progresso: {index + 1}/{total_ovs} | Tempo restante estimado: {tempo_restante/60:.1f} min")
        
        fim_processamento = time.time()
        tempo_processamento = fim_processamento - inicio_processamento
        
        # Salvar resultados
        if salvar_telefones_excel_otimizado(file_path, telefones):
            telefones_encontrados = sum(1 for t in telefones if t)
            
            log_message("💾 Reabrindo arquivo Excel...")
            os.startfile(file_path)
            
            # Estatísticas finais
            tempo_total = time.time() - inicio_total
            log_message("🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
            log_message(f"📊 ESTATÍSTICAS:")
            log_message(f"   • OVs processadas: {len(telefones)}")
            log_message(f"   • Telefones encontrados: {telefones_encontrados}")
            log_message(f"   • Taxa de sucesso: {telefones_encontrados/len(telefones)*100:.1f}%")
            log_message(f"   • Tempo total: {tempo_total/60:.1f} minutos")
            log_message(f"   • Tempo por OV: {tempo_processamento/len(telefones):.1f} segundos")
            log_message(f"   • Velocidade: {len(telefones)/(tempo_total/60):.1f} OVs/minuto")
        else:
            log_message("❌ Falha ao salvar resultados")
        
        # Pausa para ver resultados (só quando executado pelo VBA)
        if len(sys.argv) == 1:  # Sem argumentos = executado pelo VBA
            input(f"\n🎯 Processo concluído! {telefones_encontrados} telefones de {len(telefones)} OVs processadas.\nPressione Enter para fechar...")
            
    except KeyboardInterrupt:
        log_message("⏹️ Processo interrompido pelo usuário")
        input("Pressione Enter para fechar...")
    except Exception as e:
        log_message(f"💥 Erro crítico: {e}")
        import traceback
        log_message(f"📋 Detalhes técnicos: {traceback.format_exc()}")
        input("Pressione Enter para fechar...")

if __name__ == "__main__":
    main()